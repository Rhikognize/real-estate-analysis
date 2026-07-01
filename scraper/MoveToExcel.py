from openpyxl import Workbook


def create_excel():
    wb = Workbook()

    ws = wb.active
    ws.title = "Real estate data"
    headers = [
        "Title",
        "Location",
        "Rooms",
        "Shower Rooms",
        "Area",
        "Type",
        "Housing stock",
        "Price",
        "Floor",
        "Heating",
        "Has furniture",
        "Has AC",
        "Has underfloor heating",
        "Has double glazed windows",
        "Destination",
        "URL",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    return wb, ws


def save_to_excel(ws, row, data):
    for col, value in enumerate(data.values(), start=1):
        ws.cell(row=row, column=col, value=value)
